#!/usr/bin/env python3
"""Convert the AFI Tier 1 workbook into a reviewable EKB import CSV.

Rules follow the web importer in src/lib/excel-importer.ts:
- compact sheet-name matching (spaces, punctuation, 31-char truncation)
- archive product tabs included by default
- unknown product tabs imported when they have CATEGORY / SUB TIPE headers
- screenshot URLs from cells, extra columns, hyperlinks, and continuation rows
- pasted/embedded pictures are reported but not converted to screenshots
- one CSV row per scenario (Tier 1 + eskalasi stay on the same pedoman)
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import zipfile
from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


STANDARD_SHEETS = {
    "akulakupaylaterinternal": ("akulaku-internal", "Akulaku Paylater Internal"),
    "openpayonline": ("openpay-online", "Openpay Online"),
    "openpayoffline": ("openpay-offline", "Openpay Offline"),
    "akulakupaylaterditokoafia": ("paylater-toko", "Akulaku Paylater di Toko"),
    "akulakupaylaterditoko": ("paylater-toko", "Akulaku Paylater di Toko"),
    "cicilanmotorlistrik": ("cicilan-motor", "Cicilan Motor Listrik"),
    "lazadapaylater": ("lazada-paylater", "Lazada PayLater"),
    "tiktokpaylater": ("tiktok-paylater", "TikTok PayLater"),
    "general": ("general", "General"),
    "akulakuelitecard": ("elite-card", "Akulaku Elite Card"),
    "autoloan": ("auto-loan", "Auto Loan"),
}

ARCHIVE_KEYS = {"akulakuelitecard", "autoloan"}
MAPPING_SHEETS = {"rawdata", "list", "listnew"}
SPECIAL_SHEETS = {
    "ketentuanverifikasidata": "ketentuan verifikasi data",
    "logikapembuatantiket": "logika pembuatan tiket",
    "ojkspecialcase": "ojk special case",
    "transferchatcallkeasi": "transfer chatcall ke asi",
    "specialtreatmentreminder": "special treatmentreminder",
    "anomaliticketcase": "anomali ticketcase",
    "otherappcontact": "other app contact",
}

OUTPUT_COLUMNS = [
    "title",
    "product",
    "category",
    "ticket_subtype",
    "condition",
    "probing_livechat",
    "script_livechat",
    "script_callcenter",
    "images",
    "agent_steps",
    "crm_status",
    "escalation_steps",
    "escalation_status",
    "escalation_team",
    "warning",
    "resolution_type",
    "priority",
    "important_update",
    "content_status",
    "needs_review",
    "review_reason",
    "source_sheet",
    "source_type",
    "source_row",
    "source_variant",
    "duplicate_count",
]

IMAGE_HOST = re.compile(
    r"(?:drive\.google\.com|docs\.google\.com/(?:uc|file)|googleusercontent\.com|ibb\.co|imgbb\.com|\.(?:png|jpe?g|webp|gif)(?:[?#]|$))",
    re.I,
)
DOC_LINK = re.compile(r"docs\.google\.com/(?:document|spreadsheets|presentation)|drive\.google\.com/drive/folders/", re.I)
URL_RE = re.compile(r"https?://[^\s<>\"']+", re.I)
HYPERLINK_META = re.compile(r"\s*\[KORA_LINK:\s*https?://[^\]]+\]\s*", re.I)


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return re.sub(r"[ \t]+\n", "\n", text)


def compact_sheet_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", value.strip().lower())


def normalized(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


def strip_hyperlink_metadata(value: str) -> str:
    return re.sub(r"\s+", " ", HYPERLINK_META.sub(" ", clean(value))).strip()


def image_source_url(value: str) -> str:
    return re.sub(r"[)\]}]+$", "", re.sub(r"[.,;:!?]+$", "", value.strip())).strip()


def is_image_reference(value: str) -> bool:
    if DOC_LINK.search(value):
        return False
    return bool(IMAGE_HOST.search(value))


def collect_image_urls(value: str) -> list[str]:
    if not value:
        return []
    found: list[str] = []
    normalized_value = re.sub(r"https?://", lambda match: f"\n{match.group(0)}", value, flags=re.I)
    normalized_value = re.sub(r",\s*", "\n", normalized_value)
    normalized_value = re.sub(r"\s*\|\s*", "\n", normalized_value)
    for raw in URL_RE.findall(normalized_value):
        found.append(image_source_url(raw))
    for match in re.finditer(r'href=["\'](https?://[^"\']+)["\']', value, re.I):
        found.append(image_source_url(match.group(1)))
    return found


def image_identity(url: str) -> str:
    drive = re.search(r"/(?:file/)?d/([^/]+)", url, re.I)
    if drive and re.search(r"drive\.google\.com|docs\.google\.com|googleusercontent\.com", url, re.I):
        return f"drive:{drive.group(1)}"
    return url.rstrip("/")


def extract_image_references(*values: str) -> list[str]:
    seen: set[str] = set()
    images: list[str] = []
    for value in values:
        for raw in collect_image_urls(value):
            url = image_source_url(raw)
            identity = image_identity(url)
            if not is_image_reference(url) or identity in seen:
                continue
            seen.add(identity)
            images.append(url)
    return images


def cell_text(cell: Cell) -> str:
    text = clean(cell.value)
    target = ""
    hyperlink = getattr(cell, "hyperlink", None)
    if hyperlink is not None:
        target = clean(getattr(hyperlink, "target", "") or getattr(hyperlink, "location", ""))
    if target.startswith("http") and target not in text:
        text = f"{text}\n[KORA_LINK:{target}]" if text else f"[KORA_LINK:{target}]"
    return text


def row_values(row: tuple[Cell, ...], min_cols: int = 9) -> list[str]:
    values = [cell_text(cell) for cell in row]
    if len(values) < min_cols:
        values.extend([""] * (min_cols - len(values)))
    return values


def has_value(*values: str) -> bool:
    return any(value.strip() for value in values)


def has_transfer_marker(*values: str) -> bool:
    text = normalized(" ".join(values))
    return bool(re.search(r"\btf\s*asi\b", text) or ("transfer" in text and "asi" in text))


def first_meaningful_line(text: str) -> str:
    for line in text.splitlines():
        line = re.sub(r"^[\s\-–—•]+", "", line).strip()
        if line and not line.lower().startswith(("kondisi:", "status di console:", "link:")):
            return line
    return "Pedoman tanpa judul"


def inspect_embedded_drawings(path: Path) -> dict[str, int | list[str]]:
    try:
        with zipfile.ZipFile(path) as archive:
            names = archive.namelist()
    except OSError:
        return {"mediaFiles": 0, "drawingFiles": 0, "mediaNames": []}
    media = [name for name in names if name.startswith("xl/media/") and not name.endswith("/")]
    drawings = [name for name in names if re.fullmatch(r"xl/drawings/drawing\d+\.xml", name)]
    return {"mediaFiles": len(media), "drawingFiles": len(drawings), "mediaNames": media}


def match_standard_sheet(sheet_name: str) -> tuple[str, str] | None:
    compact = compact_sheet_key(sheet_name)
    if compact in STANDARD_SHEETS:
        return STANDARD_SHEETS[compact]
    prefix = [
        (key, value)
        for key, value in STANDARD_SHEETS.items()
        if len(key) >= 10 and len(compact) >= 10 and (compact.startswith(key) or key.startswith(compact))
    ]
    prefix.sort(key=lambda item: len(item[0]), reverse=True)
    return prefix[0][1] if prefix else None


def match_special_sheet(sheet_name: str) -> str | None:
    compact = compact_sheet_key(sheet_name)
    if compact in SPECIAL_SHEETS:
        return SPECIAL_SHEETS[compact]
    for key, kind in SPECIAL_SHEETS.items():
        if key in compact or compact in key:
            return kind
    return None


def is_header_like_row(row: list[str]) -> bool:
    first = normalized(row[0] if row else "")
    second = normalized(row[1] if len(row) > 1 else "")
    return bool(re.match(r"^(category|kategori)$", first) and re.match(r"^(sub\s*tipe|subtipe|sub\s*type)\b", second))


def find_standard_header_row(rows: list[list[str]]) -> int:
    for index, row in enumerate(rows[:10]):
        if is_header_like_row(row):
            return index
    return -1


def looks_like_standard_product_sheet(rows: list[list[str]]) -> bool:
    return find_standard_header_row(rows) >= 0


def worksheet_rows(ws: Worksheet) -> list[list[str]]:
    rows: list[list[str]] = []
    for row in ws.iter_rows():
        rows.append(row_values(row))
    return rows


def new_record(
    *,
    product: str,
    category: str,
    subtype: str,
    condition: str,
    live_script: str,
    call_script: str,
    images: list[str],
    agent_steps: str,
    crm_status: str,
    escalation_steps: str = "",
    escalation_status: str = "",
    escalation_team: str = "",
    warning: str = "",
    resolution: str,
    priority: str,
    source_sheet: str,
    source_row: int,
    source_variant: str,
    source_type: str,
) -> dict[str, str]:
    product = strip_hyperlink_metadata(product)
    category = strip_hyperlink_metadata(category)
    subtype = strip_hyperlink_metadata(subtype)
    condition = strip_hyperlink_metadata(condition)
    live_script = strip_hyperlink_metadata(live_script)
    call_script = strip_hyperlink_metadata(call_script)
    agent_steps = strip_hyperlink_metadata(agent_steps)
    crm_status = strip_hyperlink_metadata(crm_status)
    escalation_steps = strip_hyperlink_metadata(escalation_steps)
    escalation_status = strip_hyperlink_metadata(escalation_status)
    warning = strip_hyperlink_metadata(warning)
    if not subtype:
        subtype = "Belum dikategorikan"
    reasons: list[str] = []
    if subtype == "Belum dikategorikan":
        reasons.append("Sub tipe tiket kosong")
    if not condition:
        reasons.append("Kondisi pelanggan kosong")
    if not live_script and "Transfer" not in resolution:
        reasons.append("Skrip Live Chat kosong")
    if not agent_steps and not escalation_steps and "Referensi" not in resolution and "Transfer" not in resolution:
        reasons.append("Langkah agent kosong")

    return {
        "title": first_meaningful_line(condition or subtype),
        "product": product,
        "category": category or "Belum dikategorikan",
        "ticket_subtype": subtype,
        "condition": condition or subtype or "Kondisi pelanggan belum diisi",
        "probing_livechat": condition,
        "script_livechat": live_script or "Skrip Live Chat belum diisi pada sumber.",
        "script_callcenter": call_script,
        "images": " | ".join(images),
        "agent_steps": agent_steps,
        "crm_status": crm_status,
        "escalation_steps": escalation_steps,
        "escalation_status": escalation_status,
        "escalation_team": escalation_team,
        "warning": warning,
        "resolution_type": resolution,
        "priority": priority,
        "important_update": "Tidak",
        "content_status": "Published",
        "needs_review": "Ya" if reasons else "Tidak",
        "review_reason": "; ".join(dict.fromkeys(reasons)),
        "source_sheet": source_sheet,
        "source_row": str(source_row),
        "source_variant": source_variant,
        "source_type": source_type,
        "duplicate_count": "1",
    }


def merge_images(first: str, second: list[str]) -> str:
    existing = [part.strip() for part in first.split("|") if part.strip()]
    return " | ".join(extract_image_references(*(existing + second)))


def read_standard_sheet(sheet_name: str, rows: list[list[str]], product: str, source_type: str) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    header_row = find_standard_header_row(rows)
    start = header_row + 1 if header_row >= 0 else 3
    category = ""
    subtype = ""
    for index in range(start, len(rows)):
        row = rows[index]
        if not any(row) or is_header_like_row(row):
            continue
        category = row[0] or category
        subtype = row[1] or subtype
        images = extract_image_references(*row)
        if not has_value(*row[2:9]):
            if images and records:
                records[-1]["images"] = merge_images(records[-1]["images"], images)
            continue
        force_transfer = has_transfer_marker(row[2], row[3], row[5], row[6])
        resolution = "Transfer ke ASI" if force_transfer else "Selesai di Tier 1"
        if has_value(row[7], row[8]) and not force_transfer:
            resolution = "Selesai di Tier 1 + Eskalasi Tier 2/3" if has_value(row[5], row[6]) else "Eskalasi Tier 2/3"
        records.append(
            new_record(
                product=product,
                category=category,
                subtype=subtype,
                condition=row[2],
                live_script=row[3],
                call_script=row[4],
                images=images,
                agent_steps=row[5],
                crm_status=row[6],
                escalation_steps=row[7],
                escalation_status=row[8],
                warning="",
                resolution=resolution,
                priority="Normal",
                source_sheet=sheet_name,
                source_row=index + 1,
                source_variant="scenario",
                source_type=source_type,
            )
        )
    return records


def read_special_sheet(sheet_name: str, rows: list[list[str]], kind: str) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    if kind == "ketentuan verifikasi data":
        for index in range(2, len(rows)):
            row = rows[index]
            if not any(row):
                continue
            livechat = row[1] or row[0]
            records.append(
                new_record(
                    product="Pedoman Operasional",
                    category="Ketentuan Verifikasi Data",
                    subtype=f"Verifikasi data {index - 1}",
                    condition=livechat,
                    live_script=livechat,
                    call_script=row[0],
                    images=extract_image_references(*row),
                    agent_steps=livechat,
                    crm_status="Kategori Percakapan",
                    resolution="Selesai di Tier 1",
                    priority="Normal",
                    source_sheet=sheet_name,
                    source_row=index + 1,
                    source_variant="livechat",
                    source_type="operational_verification",
                )
            )
    elif kind == "logika pembuatan tiket":
        section = "Logika Tiket"
        for index in range(2, len(rows)):
            row = rows[index]
            content = next((value for value in row if value), "")
            if not content:
                continue
            if len(content) < 90 and not re.match(r"^\d+[.)]", content):
                section = content
                continue
            records.append(
                new_record(
                    product="Pedoman Operasional",
                    category="Logika Pembuatan Tiket",
                    subtype=section,
                    condition=content,
                    live_script="",
                    call_script="",
                    images=extract_image_references(*row),
                    agent_steps=content,
                    crm_status="Kategori Percakapan",
                    resolution="Selesai di Tier 1",
                    priority="Normal",
                    source_sheet=sheet_name,
                    source_row=index + 1,
                    source_variant="tier1",
                    source_type="operational_ticket_logic",
                )
            )
    elif kind == "ojk special case":
        for index in range(2, len(rows)):
            row = rows[index]
            if not any(row):
                continue
            records.append(
                new_record(
                    product="Pedoman Operasional",
                    category="OJK Special Case",
                    subtype=row[1],
                    condition=row[2],
                    live_script=row[3],
                    call_script=row[4],
                    images=extract_image_references(*row),
                    agent_steps=row[5],
                    crm_status=row[6],
                    escalation_steps=row[7],
                    escalation_status=row[8],
                    resolution="Special Case",
                    priority="Special Case",
                    source_sheet=sheet_name,
                    source_row=index + 1,
                    source_variant="scenario",
                    source_type="special_ojk",
                )
            )
    elif kind == "transfer chatcall ke asi":
        for index in range(2, len(rows)):
            row = rows[index]
            if not any(row):
                continue
            records.append(
                new_record(
                    product="Pedoman Operasional",
                    category="Transfer Chat/Call ke ASI",
                    subtype="Transfer chat/call ke ASI",
                    condition=row[0],
                    live_script=row[4],
                    call_script=row[5],
                    images=extract_image_references(*row),
                    agent_steps=row[1],
                    crm_status="",
                    warning=row[2],
                    escalation_team="ASI",
                    resolution="Transfer ke ASI",
                    priority="Low",
                    source_sheet=sheet_name,
                    source_row=index + 1,
                    source_variant="transfer",
                    source_type="special_transfer",
                )
            )
    elif kind == "special treatmentreminder":
        for index in range(2, len(rows)):
            row = rows[index]
            if not any(row):
                continue
            records.append(
                new_record(
                    product="Pedoman Operasional",
                    category="Special Treatment / Reminder",
                    subtype="Special Treatment",
                    condition=row[0],
                    live_script="",
                    call_script="",
                    images=extract_image_references(*row),
                    agent_steps="",
                    crm_status="",
                    escalation_steps=row[1],
                    warning="\n\n".join(value for value in (row[3], row[4]) if value),
                    escalation_team=row[2],
                    resolution="Eskalasi Tier 2/3",
                    priority="Special Case",
                    source_sheet=sheet_name,
                    source_row=index + 1,
                    source_variant="escalation",
                    source_type="special_treatment",
                )
            )
    elif kind == "anomali ticketcase":
        for index in range(1, len(rows)):
            row = rows[index]
            if not any(row):
                continue
            records.append(
                new_record(
                    product="Pedoman Operasional",
                    category="Anomali Ticket/Case",
                    subtype=f"Anomali {row[0]}" if row[0] else "Anomali Ticket/Case",
                    condition=row[1],
                    live_script="",
                    call_script="",
                    images=extract_image_references(*row),
                    agent_steps=row[2],
                    crm_status="",
                    resolution="Referensi saja",
                    priority="Special Case",
                    source_sheet=sheet_name,
                    source_row=index + 1,
                    source_variant="reference",
                    source_type="special_anomaly",
                )
            )
    elif kind == "other app contact":
        for index in range(1, len(rows)):
            row = rows[index]
            if not any(row):
                continue
            platform = row[0] or "Platform lain"
            records.append(
                new_record(
                    product="Pedoman Operasional",
                    category="Other App Contact",
                    subtype=f"Kontak {platform}",
                    condition=f"Informasi kontak {platform}",
                    live_script=row[1],
                    call_script="",
                    images=extract_image_references(*row),
                    agent_steps="",
                    crm_status="",
                    resolution="Referensi saja",
                    priority="Normal",
                    source_sheet=sheet_name,
                    source_row=index + 1,
                    source_variant="reference",
                    source_type="other_contact",
                )
            )
            records[-1]["needs_review"] = "Tidak"
            records[-1]["review_reason"] = ""
    return records


def deduplicate(records: Iterable[dict[str, str]]) -> list[dict[str, str]]:
    groups: defaultdict[tuple[str, ...], list[dict[str, str]]] = defaultdict(list)
    key_fields = ("product", "category", "ticket_subtype", "condition", "script_livechat", "resolution_type")
    for record in records:
        groups[tuple(normalized(record[field]) for field in key_fields)].append(record)

    output: list[dict[str, str]] = []
    for group in groups.values():
        record = dict(group[0])
        record["duplicate_count"] = str(len(group))
        images: list[str] = []
        for item in group:
            images.extend(part.strip() for part in item["images"].split("|") if part.strip())
        record["images"] = " | ".join(extract_image_references(*images))
        if len(group) > 1:
            record["needs_review"] = "Ya"
            existing = [part for part in (record["review_reason"], "Duplikat persis sumber") if part]
            record["review_reason"] = "; ".join(existing)
        output.append(record)
    return output


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path, help="Path workbook AFI .xlsx")
    parser.add_argument("--output", required=True, type=Path, help="Path CSV hasil bulk import")
    parser.add_argument("--exclude-archived", action="store_true", help="Lewati sheet produk arsip (Elite Card / Auto Loan)")
    parser.add_argument("--include-archived", action="store_true", help=argparse.SUPPRESS)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    drawings = inspect_embedded_drawings(args.input)
    workbook = load_workbook(args.input, read_only=False, data_only=True)
    records: list[dict[str, str]] = []
    skipped: list[str] = []
    imported_tabs: list[str] = []

    for sheet_name in workbook.sheetnames:
        compact = compact_sheet_key(sheet_name)
        ws = workbook[sheet_name]
        rows = worksheet_rows(ws)
        if compact == "listnew":
            continue
        if compact in MAPPING_SHEETS:
            continue
        special = match_special_sheet(sheet_name)
        if special:
            parsed = read_special_sheet(sheet_name, rows, special)
            records.extend(parsed)
            imported_tabs.append(sheet_name)
            continue
        standard = match_standard_sheet(sheet_name)
        if standard:
            _product_id, product = standard
            source_type = "archived" if compact_sheet_key(sheet_name) in ARCHIVE_KEYS or compact in ARCHIVE_KEYS else "standard"
            if source_type == "archived" and args.exclude_archived:
                skipped.append(sheet_name)
                continue
            records.extend(read_standard_sheet(sheet_name, rows, product, source_type))
            imported_tabs.append(sheet_name)
            continue
        if looks_like_standard_product_sheet(rows):
            product = re.sub(r"\s+", " ", sheet_name).strip()
            records.extend(read_standard_sheet(sheet_name, rows, product, "standard"))
            imported_tabs.append(sheet_name)
            continue
        skipped.append(sheet_name)

    records = deduplicate(records)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)

    report = {
        "input": str(args.input),
        "output": str(args.output),
        "rows_output": len(records),
        "rows_needing_review": sum(record["needs_review"] == "Ya" for record in records),
        "duplicate_rows_collapsed": sum(max(int(record["duplicate_count"]) - 1, 0) for record in records),
        "with_screenshots": sum(bool(record["images"]) for record in records),
        "by_source_type": dict(Counter(record["source_type"] for record in records)),
        "by_resolution": dict(Counter(record["resolution_type"] for record in records)),
        "imported_tabs": imported_tabs,
        "skipped_sheets": skipped,
        "archives_included": not args.exclude_archived,
        "embedded_drawings": drawings,
        "pasted_image_note": (
            "Gambar tempel di xl/media tidak masuk CSV. Letakkan URL Drive/ibb di sel."
            if drawings["mediaFiles"]
            else "Tidak ada gambar tempel; screenshot mengikuti URL di sel."
        ),
    }
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
